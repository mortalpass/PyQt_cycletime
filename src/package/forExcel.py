from pathlib import Path
from typing import List, Optional
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter
import logging
import json

# 配置日志
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    datefmt='%Y-%m-%d %H:%M:%S'
)
logger = logging.getLogger(__name__)


class ExcelExporter:
    """Excel导出器类，用于将CSV文件转换为格式化的Excel工作簿"""

    # 定义样式常量
    RED_FILL = PatternFill(start_color="FFFF0000", end_color="FFFF0000", fill_type="solid")
    CAMBRIA_FONT = Font(name='Cambria', size=14)
    BOLD_CAMBRIA_FONT = Font(name='Cambria', size=14, bold=True)
    CENTER_ALIGNMENT = Alignment(horizontal='left', vertical='center')

    def __init__(self, form_dir: str, sep: str = ';'):
        self.form_dir = Path(form_dir)
        self.sep = sep

    def find_workstation_dirs(self) -> List[Path]:
        """查找所有工作站目录"""
        workstation_dirs = []
        try:
            for item in self.form_dir.iterdir():
                if item.is_dir() and not item.name.startswith('.'):
                    workstation_dirs.append(item)
            logger.info(f"找到 {len(workstation_dirs)} 个工作站目录")
        except Exception as e:
            logger.error(f"查找工作站目录时出错: {e}")
        return workstation_dirs

    def find_comparison_files(self, workstation_dir: Path) -> List[Path]:
        """在工作站目录中查找比较文件"""
        comparison_files = []
        try:
            # 查找所有以_comparison.csv结尾的文件
            pattern = "*_comparison.csv"
            found_files = list(workstation_dir.glob(pattern))
            comparison_files.extend(found_files)
            logger.info(f"在 {workstation_dir.name} 中找到 {len(comparison_files)} 个比较文件")
        except Exception as e:
            logger.error(f"在 {workstation_dir.name} 中查找比较文件时出错: {e}")
        return comparison_files

    def _update_metadata_paths(self, metadata_file: str, excel_path: Path) -> None:
        """
        更新metadata.json文件中的路径信息

        参数:
            metadata_file: metadata.json文件路径
            excel_path: 生成的Excel文件路径
        """
        try:
            # 读取现有的metadata.json文件
            with open(metadata_file, 'r', encoding='utf-8') as f:
                metadata = json.load(f)

            # 确保最外层有path字段
            if 'path' not in metadata:
                metadata['path'] = {}

            # 更新路径信息
            metadata['path']['form'] = str(excel_path)
            metadata['path']['json'] = str(metadata_file)

            # 写回文件
            with open(metadata_file, 'w', encoding='utf-8') as f:
                json.dump(metadata, f, indent=2, ensure_ascii=False)

            logger.info(f"已更新metadata.json中的路径信息: {metadata_file}")

        except Exception as e:
            logger.error(f"更新metadata.json路径时出错: {e}")

    def export_all_workstations_to_excel(
            self,
            output_path: str = None,
            metadata_file: str = None
    ) -> bool:
        """
        将所有工作站的比较文件导出到Excel工作簿的不同工作表

        参数:
            output_path: 输出的Excel文件完整路径
            metadata_file: metadata.json文件路径，用于创建metadata工作表

        返回:
            bool: 操作是否成功
        """
        # 查找所有工作站目录
        workstation_dirs = self.find_workstation_dirs()
        if not workstation_dirs:
            logger.error("未找到任何工作站目录")
            return False

        output_path = Path(output_path) if output_path else self.form_dir / "all_comparisons.xlsx"

        try:
            # 创建Excel文件
            with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
                success_count = 0

                # 处理每个工作站
                for workstation_dir in workstation_dirs:
                    # 查找比较文件
                    comparison_files = self.find_comparison_files(workstation_dir)

                    for csv_file in comparison_files:
                        try:
                            # 使用工作站名称作为工作表名
                            sheet_name = workstation_dir.name

                            # 如果工作表名已存在，添加后缀
                            if sheet_name in writer.book.sheetnames:
                                sheet_name = f"{workstation_dir.name}_{success_count + 1}"

                            # 读取CSV并写入Excel
                            df = pd.read_csv(csv_file, sep=self.sep, encoding='utf-8')
                            df.to_excel(writer, sheet_name=sheet_name, index=False)
                            logger.info(f"已添加工作表: {sheet_name}")
                            success_count += 1
                        except Exception as e:
                            logger.error(f"处理文件 {csv_file} 时出错: {e}")
                            continue

                # 检查是否成功添加了至少一个工作表
                if success_count == 0:
                    logger.error("未能添加任何工作表到Excel文件")
                    return False

                # 如果提供了metadata文件，添加metadata工作表
                if metadata_file:
                    self._add_metadata_sheet(writer, metadata_file)

            # 应用格式设置
            if output_path.exists():
                self._apply_excel_formatting(output_path)
                logger.info(f"格式设置完成! 结果已保存到 {output_path}")

                # 新增：更新metadata.json中的路径信息
                if metadata_file:
                    self._update_metadata_paths(metadata_file, output_path)

                return True
            else:
                logger.error(f"输出文件 {output_path} 未创建")
                return False

        except Exception as e:
            logger.error(f"导出Excel时发生错误: {e}")
            return False

    def _add_metadata_sheet(self, writer, metadata_file: str) -> None:
        """
        从metadata.json文件创建metadata工作表，适应新的JSON格式

        参数:
            writer: ExcelWriter对象
            metadata_file: metadata.json文件路径
        """
        try:
            # 读取JSON文件
            with open(metadata_file, 'r', encoding='utf-8') as f:
                metadata = json.load(f)

            # 创建metadata工作表
            workbook = writer.book
            if "Metadata" in workbook.sheetnames:
                del workbook["Metadata"]
            ws = workbook.create_sheet("Metadata")

            # 写入metadata部分
            ws.append(["Field", "Description"])

            # 检查JSON结构并写入相应的字段
            if "metadata_dict" in metadata:
                for key, value in metadata["metadata_dict"].items():
                    ws.append([key, value])
            else:
                # 如果没有metadata_dict，尝试直接写入顶级字段
                for key, value in metadata.items():
                    if key != "path" and not key.startswith("WS-"):
                        ws.append([key, value])

            # 添加空行
            for _ in range(3):
                ws.append([])

            # 找出所有工作站键 (WS-1, WS-2等)
            workstation_keys = [key for key in metadata.keys() if key.startswith("WS-")]

            if not workstation_keys:
                logger.warning("未找到工作站数据，跳过数据表创建")
                return

            # 获取第一个工作站的所有字段作为表头，但排除path字段
            first_ws = workstation_keys[0]
            headers = ["Workstation ID"] + [key for key in metadata[first_ws].keys() if key != "path"]
            ws.append(headers)

            # 为每个工作站添加一行数据，排除path字段
            for ws_key in workstation_keys:
                ws_data = metadata[ws_key]
                row_data = [ws_key]  # 工作站ID

                # 按表头顺序添加数据，排除path字段
                for header in headers[1:]:  # 跳过第一个"Workstation ID"
                    row_data.append(ws_data.get(header, ""))

                ws.append(row_data)

            # 添加备注行
            if "note" in metadata:
                ws.append([])
                ws.append([metadata["note"]])

            # 设置Metadata工作表的格式
            self._format_metadata_sheet(ws)

            logger.info("已添加Metadata工作表")

        except Exception as e:
            logger.error(f"创建metadata工作表时出错: {e}")

    def _format_metadata_sheet(self, worksheet) -> None:
        """设置Metadata工作表的格式"""
        # 设置所有单元格的字体为Cambria，大小为14
        for row in worksheet.iter_rows(min_row=1, max_row=worksheet.max_row, min_col=1, max_col=worksheet.max_column):
            for cell in row:
                if cell.row == 1:  # 标题行
                    cell.font = self.BOLD_CAMBRIA_FONT
                else:  # 数据行
                    cell.font = self.CAMBRIA_FONT
                cell.alignment = self.CENTER_ALIGNMENT

        # 调整列宽
        self._adjust_column_widths(worksheet)

    def _apply_excel_formatting(self, excel_path: Path) -> None:
        """
        应用Excel格式设置：标记时间差大于1的单元格并调整列宽

        参数:
            excel_path: Excel文件路径
        """
        try:
            # 加载工作簿
            wb = load_workbook(excel_path)

            # 处理每个工作表
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]

                # 设置行高为22
                for row in range(1, ws.max_row + 1):
                    ws.row_dimensions[row].height = 22

                # 如果不是Metadata工作表，查找Time_Difference列
                if sheet_name != "Metadata":
                    time_diff_col = self._find_column_index(ws, "Time_Difference")

                    if time_diff_col:
                        # 设置标题格式
                        header_cell = ws.cell(row=1, column=time_diff_col)
                        header_cell.font = self.BOLD_CAMBRIA_FONT
                        header_cell.alignment = self.CENTER_ALIGNMENT

                        # 标记大于1的单元格
                        self._highlight_cells_above_threshold(ws, time_diff_col, 1.0)

                # 设置所有单元格的字体和居中对齐
                for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
                    for cell in row:
                        if cell.row == 1:  # 标题行
                            cell.font = self.BOLD_CAMBRIA_FONT
                        else:  # 数据行
                            cell.font = self.CAMBRIA_FONT
                        cell.alignment = self.CENTER_ALIGNMENT

                # 调整列宽
                self._adjust_column_widths(ws)

            # 保存修改
            wb.save(excel_path)

        except Exception as e:
            logger.error(f"应用Excel格式时出错: {e}")

    def _find_column_index(self, worksheet, column_name: str) -> Optional[int]:
        """查找指定列名的列索引"""
        for col_idx in range(1, worksheet.max_column + 1):
            cell_value = worksheet.cell(row=1, column=col_idx).value
            if cell_value and column_name.lower() in str(cell_value).lower():
                return col_idx
        return None

    def _highlight_cells_above_threshold(self, worksheet, column_index: int, threshold: float) -> None:
        """标记大于阈值的单元格"""
        for row in range(2, worksheet.max_row + 1):
            cell = worksheet.cell(row=row, column=column_index)
            try:
                if cell.value and float(cell.value) > threshold:
                    cell.fill = self.RED_FILL
            except (ValueError, TypeError):
                # 忽略无法转换为浮点数的单元格
                continue

    def _adjust_column_widths(self, worksheet) -> None:
        """自适应调整所有列宽"""
        for col in worksheet.columns:
            max_length = 0
            column_letter = get_column_letter(col[0].column)

            # 查找最大内容长度
            for cell in col:
                try:
                    # 考虑标题行可能需要额外宽度
                    is_header = cell.row == 1
                    cell_length = len(str(cell.value)) + (2 if is_header else 0)
                    max_length = max(max_length, cell_length)
                except Exception:
                    continue

            # 设置列宽（限制最大宽度）
            adjusted_width = min(max_length + 2, 50)
            worksheet.column_dimensions[column_letter].width = adjusted_width


def main():
    """主函数"""
    try:
        from config import output_form

        # 创建导出器实例
        exporter = ExcelExporter(form_dir=output_form, sep=';')

        # 输出Excel文件路径
        output_excel = Path(output_form) / 'all_comparisons.xlsx'
        metadata_file = Path(output_form) / 'metadata.json'

        # 执行导出
        success = exporter.export_all_workstations_to_excel(
            output_path=output_excel,
            metadata_file=metadata_file
        )

        if success:
            logger.info("Excel导出完成!")
        else:
            logger.error("Excel导出失败!")

    except ImportError:
        logger.error("无法导入配置模块，请确保 config.py 存在")
    except Exception as e:
        logger.error(f"处理过程中发生错误: {e}")


if __name__ == "__main__":
    main()